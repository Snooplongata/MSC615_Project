# -*- coding: utf-8 -*-
"""
Created on Wed Nov  7 19:40:44 2018

@author: Marcel
"""
import pulp
import numpy as np
from itertools import permutations
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side

def get_sample(arr, n_iter=None, sample_size = 10, fast = True):
    """
    
    """
    n = len(arr)
    if fast:
        start_idx = (n_iter * sample_size) % n
        if start_idx + sample_size >= n:
            np.random.shuffle(arr)
        return arr[start_idx:start_idx+sample_size]
    else:
        return np.random.choice(arr, sample_size, replace = False)

def cost_func(tc, places, sample):       
    costs = []
    for perm_i in sample:
        total = 0
        last = 0
        for j in perm_i[1:]:
            if not last:
                total += tc["Dominoe's Pizza"][places[j-1]]
                last = j-1
            else:
                total += tc[places[last]][places[j-1]]
                last = j-1
        
        costs.append(np.hstack((total,perm_i)))  
    
    costs = np.array(costs)
    return costs

def create_model(places, samples, z = -1):
    wb = Workbook()
    ws = wb.active
    
    ws.title = "Main Model"
    
    prov_fill = PatternFill("solid", fgColor = "00B0F0")
    change_fill = PatternFill("solid", fgColor = "FFFF00")
    of_fill = PatternFill("solid", fgColor = "FFC000")
    border = Border(top = openpyxl.styles.borders.Side(style = 'thin'), 
                 bottom = openpyxl.styles.borders.Side(style = 'thin'), 
                 left = openpyxl.styles.borders.Side(style = 'thin'), 
                 right = openpyxl.styles.borders.Side(style = 'thin'))
    
    n_routes = len(samples)
    
    costs_label = ws.cell(row = 2, column = 1, value = "Costs")
    costs = "{}2:{}2".format(chr(66),chr(66+n_routes-1))
    for i in range(n_routes):
        cost = samples[i,0]
        ws.cell(row = 2, column = 2 + i, value = cost)
        ws[chr(66+i)+"2"].fill = prov_fill
    
    routes_label = ws.cell(row = 4, column = 1, value = "Routes")
    changingroutes = "{}4:{}4".format(chr(66),chr(66+n_routes-1))
    for i in range(n_routes):
        ws.cell(row = 4, column = 2 + i, value = 0)
        ws[chr(66+i)+"4"].fill = change_fill
    ws.cell(row = 4, column = n_routes+2, value = "=SUM({})".format(changingroutes))
    ws[chr(66+n_routes)+"4"].border = border
    ws.cell(row = 4, column = n_routes+3, value = "=")
    ws.cell(row = 4, column = n_routes+4, value = 6)
    ws[chr(68+n_routes)+"4"].fill = prov_fill

    #Total Cost
    Total_label = ws.cell(row = 1, column = n_routes+4, value = "Total Cost")
    ws[chr(68+n_routes)+"2"] = "=SUMPRODUCT({},{})".format(costs,changingroutes)
    ws[chr(68+n_routes)+"2"].fill = of_fill
    ws[chr(68+n_routes)+"2"].border = border

    #Destinations
    destinationss_label = ws.cell(row = 6, column = 1, value = "Destinations")
    for i in range(len(places)):
        ws.cell(row = 7+i, column = 1, value = places[i])
        this_row = "{}{}:{}{}".format(chr(66),7+i,chr(66+n_routes-1),7+i)
        #Output for this row
        ws.cell(row = 7+i, column = n_routes+2, value = "=SUMPRODUCT({},{})".format(this_row,changingroutes))
        ws[chr(66+n_routes)+"{}".format(7+i)].border = border
        ws.cell(row = 7+i, column = n_routes+3, value = ">=")
        ws.cell(row = 7+i, column = n_routes+4, value = 1)
        ws[chr(68+n_routes)+"{}".format(7+i)].fill = prov_fill

    for i,sample in enumerate(samples):
        for j,value in enumerate(sample[1:]):
            ws.cell(row = value+6, column = 2+i, value = 1)
    
    if z == -1:
        wb.save("Model.xlsx")
    else:
        wb.save("Model_{}.xlsx")

def test_pulp(sample):
    sample = sample.astype(int)
    # create the LP object, set up as a minimization problem
    prob = pulp.LpProblem('Delivery', pulp.LpMinimize)
    
    # set up decision variables
    route1 = pulp.LpVariable('route1', lowBound=0, upBound=1, cat='Integer')
    route2 = pulp.LpVariable('route2', lowBound=0, upBound=1, cat='Integer')
    route3 = pulp.LpVariable('route3', lowBound=0, upBound=1, cat='Integer')
    route4 = pulp.LpVariable('route4', lowBound=0, upBound=1, cat='Integer')
    route5 = pulp.LpVariable('route5', lowBound=0, upBound=1, cat='Integer')
    route6 = pulp.LpVariable('route6', lowBound=0, upBound=1, cat='Integer')
    route7 = pulp.LpVariable('route7', lowBound=0, upBound=1, cat='Integer')
    route8 = pulp.LpVariable('route8', lowBound=0, upBound=1, cat='Integer')
    route9 = pulp.LpVariable('route9', lowBound=0, upBound=1, cat='Integer')
    route10 = pulp.LpVariable('route10', lowBound=0, upBound=1, cat='Integer')
    
    selection = np.array([route1,route2,route3,route4,route5,route6,route7,
                          route8,route9,route10])
    costs = sample[:,0]
    total_cost = sum(costs*selection)
    
    bip = np.zeros((30,len(costs)))
    for j,route in enumerate(sample):
        for stop in route[1:]:
            bip[stop-1,j] = 1
    
    #Objective Function
    prob += total_cost
    
    #Add constraints
    constraint1 = sum(selection)
    prob += (constraint1 == 3)
    contraint2 = sum(bip[0] * selection)
    prob += (contraint2 >= 1)
    contraint3 = sum(bip[1] * selection)
    prob += (contraint3 >= 1)
    contraint4 = sum(bip[2] * selection)
    prob += (contraint4 >= 1)
    contraint5 = sum(bip[3] * selection)
    prob += (contraint5 >= 1)
    contraint6 = sum(bip[4] * selection)
    prob += (contraint6 >= 1)
    contraint7 = sum(bip[5] * selection)
    prob += (contraint7 >= 1)
    contraint8 = sum(bip[6] * selection)
    prob += (contraint8 >= 1)
    contraint9 = sum(bip[7] * selection)
    prob += (contraint9 >= 1)
    contraint10 = sum(bip[8] * selection)
    prob += (contraint10 >= 1)
    constraint11 = sum(bip[9] * selection)
    prob += (constraint11 == 3)
    contraint12 = sum(bip[10] * selection)
    prob += (contraint12 >= 1)
    contraint13 = sum(bip[11] * selection)
    prob += (contraint13 >= 1)
    contraint14 = sum(bip[12] * selection)
    prob += (contraint14 >= 1)
    contraint15 = sum(bip[13] * selection)
    prob += (contraint15 >= 1)
    contraint16 = sum(bip[14] * selection)
    prob += (contraint16 >= 1)
    contraint17 = sum(bip[15] * selection)
    prob += (contraint17 >= 1)
    contraint18 = sum(bip[16] * selection)
    prob += (contraint18 >= 1)
    contraint19 = sum(bip[17] * selection)
    prob += (contraint19 >= 1)
    contraint20 = sum(bip[18] * selection)
    prob += (contraint20 >= 1)
    constraint21 = sum(bip[19] * selection)
    prob += (constraint21 == 3)
    contraint22 = sum(bip[20] * selection)
    prob += (contraint22 >= 1)
    contraint23 = sum(bip[21] * selection)
    prob += (contraint23 >= 1)
    contraint24 = sum(bip[22] * selection)
    prob += (contraint24 >= 1)
    contraint25 = sum(bip[23] * selection)
    prob += (contraint25 >= 1)
    contraint26 = sum(bip[24] * selection)
    prob += (contraint26 >= 1)
    contraint27 = sum(bip[25] * selection)
    prob += (contraint27 >= 1)
    contraint28 = sum(bip[26] * selection)
    prob += (contraint28 >= 1)
    contraint29 = sum(bip[27] * selection)
    prob += (contraint29 >= 1)
    contraint30 = sum(bip[28] * selection)
    prob += (contraint30 >= 1)
    contraint31 = sum(bip[29] * selection)
    prob += (contraint31 >= 1)
    
    #print(prob)
    
    # solve the LP using the default solver
    optimization_result = prob.solve()
    
    # make sure we got an optimal solution
    if optimization_result == pulp.LpStatusOptimal:    
        # display the results
        for i,var in enumerate(selection):
            print('{} was chosen: {}'.format(var.name, bool(var.value())))
            
        return True
    else:
        return False
    
#test_pulp(sample)